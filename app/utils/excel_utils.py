"""
Excel处理工具
"""
import os
from datetime import datetime
from typing import Optional
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from motor.motor_asyncio import AsyncIOMotorDatabase


async def export_candidate_scores(
    db: AsyncIOMotorDatabase,
    session_id: str,
    include_feedbacks: bool = True
) -> str:
    """
    导出候选人得分明细
    
    Args:
        db: 数据库连接
        session_id: 场次ID
        include_feedbacks: 是否包含评语
        
    Returns:
        文件路径
    """
    # 获取场次信息
    try:
        session_oid = ObjectId(session_id)
    except Exception:
        raise ValueError("无效的场次ID")
    
    session = await db.sessions.find_one({"_id": session_oid})
    if not session:
        raise ValueError("场次不存在")
    
    # 获取候选人列表
    candidates = await db.candidates.find(
        {"session_id": session_id}
    ).sort("created_at", 1).to_list(None)
    
    # 获取所有评分
    scores = await db.scores.find(
        {"session_id": session_id}
    ).to_list(None)
    
    # 获取评委信息
    interviewer_ids = list(set(score["interviewer_id"] for score in scores))
    interviewers = await db.users.find(
        {"_id": {"$in": interviewer_ids}}
    ).to_list(None)
    interviewer_map = {str(i["_id"]): i["name"] for i in interviewers}
    
    # 组织数据
    score_map = {}  # {candidate_id: {interviewer_id: score_data}}
    for score in scores:
        cid = score["candidate_id"]
        iid = score["interviewer_id"]
        if cid not in score_map:
            score_map[cid] = {}
        score_map[cid][iid] = score
    
    # 创建工作簿
    wb = _create_workbook()
    ws = wb.active
    ws.title = "候选人得分明细"
    
    # 构建表头
    headers = ["序号", "姓名", "手机号", "平均分"]
    for iid in interviewer_ids:
        headers.append(interviewer_map.get(iid, f"评委{iid[:4]}"))
    if include_feedbacks:
        headers.append("综合评语")
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # 写入数据
    for row_idx, candidate in enumerate(candidates, 2):
        cid = str(candidate["_id"])
        candidate_scores = score_map.get(cid, {})
        
        # 计算平均分
        total_scores = [s["total_score"] for s in candidate_scores.values()]
        avg_score = round(sum(total_scores) / len(total_scores), 1) if total_scores else 0
        
        # 序号、姓名、手机号、平均分
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=candidate["name"])
        ws.cell(row=row_idx, column=3, value=candidate["phone"])
        ws.cell(row=row_idx, column=4, value=avg_score)
        
        # 各评委评分
        col_idx = 5
        for iid in interviewer_ids:
            score_data = candidate_scores.get(iid)
            score_value = score_data["total_score"] if score_data else "-"
            ws.cell(row=row_idx, column=col_idx, value=score_value)
            col_idx += 1
        
        # 综合评语
        if include_feedbacks:
            feedbacks = []
            for iid in interviewer_ids:
                score_data = candidate_scores.get(iid)
                if score_data and score_data.get("feedback"):
                    interviewer_name = interviewer_map.get(iid, "评委")
                    feedbacks.append(f"{interviewer_name}: {score_data['feedback']}")
            feedback_text = "\n".join(feedbacks) if feedbacks else "-"
            ws.cell(row=row_idx, column=col_idx, value=feedback_text)
    
    # 设置列宽
    _set_column_width(ws, headers, include_feedbacks)
    
    # 应用样式
    _apply_styles(ws, len(candidates) + 1, len(headers))
    
    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"candidate_scores_{session['name']}_{timestamp}.xlsx"
    filepath = os.path.join("uploads", "exports", filename)
    
    # 确保目录存在
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath


async def export_interviewer_stats(
    db: AsyncIOMotorDatabase,
    session_id: str
) -> str:
    """
    导出评委统计
    
    Args:
        db: 数据库连接
        session_id: 场次ID
        
    Returns:
        文件路径
    """
    # 获取场次信息
    try:
        session_oid = ObjectId(session_id)
    except Exception:
        raise ValueError("无效的场次ID")
    
    session = await db.sessions.find_one({"_id": session_oid})
    if not session:
        raise ValueError("场次不存在")
    
    # 获取所有评分
    scores = await db.scores.find(
        {"session_id": session_id}
    ).to_list(None)
    
    # 获取评委信息
    interviewer_ids = list(set(score["interviewer_id"] for score in scores))
    interviewers = await db.users.find(
        {"_id": {"$in": interviewer_ids}}
    ).to_list(None)
    interviewer_map = {str(i["_id"]): i["name"] for i in interviewers}
    
    # 统计每个评委的数据
    stats_data = []
    for iid in interviewer_ids:
        interviewer_scores = [s for s in scores if s["interviewer_id"] == iid]
        
        if not interviewer_scores:
            continue
        
        total_scores = [s["total_score"] for s in interviewer_scores]
        count = len(total_scores)
        avg_score = sum(total_scores) / count
        
        # 计算标准差
        variance = sum((x - avg_score) ** 2 for x in total_scores) / count
        std_dev = variance ** 0.5
        
        # 计算极端分数量（低于60或高于95）
        extreme_count = sum(1 for s in total_scores if s < 60 or s > 95)
        
        stats_data.append({
            "name": interviewer_map.get(iid, f"评委{iid[:4]}"),
            "count": count,
            "avg_score": round(avg_score, 1),
            "std_dev": round(std_dev, 1),
            "extreme_count": extreme_count
        })
    
    # 按完成数排序
    stats_data.sort(key=lambda x: x["count"], reverse=True)
    
    # 创建工作簿
    wb = _create_workbook()
    ws = wb.active
    ws.title = "评委统计"
    
    # 表头
    headers = ["评委姓名", "完成数", "平均分", "标准差", "极端分数量"]
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # 写入数据
    for row_idx, stat in enumerate(stats_data, 2):
        ws.cell(row=row_idx, column=1, value=stat["name"])
        ws.cell(row=row_idx, column=2, value=stat["count"])
        ws.cell(row=row_idx, column=3, value=stat["avg_score"])
        ws.cell(row=row_idx, column=4, value=stat["std_dev"])
        ws.cell(row=row_idx, column=5, value=stat["extreme_count"])
    
    # 设置列宽
    column_widths = [15, 10, 10, 10, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 应用样式
    _apply_styles(ws, len(stats_data) + 1, len(headers))
    
    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"interviewer_stats_{session['name']}_{timestamp}.xlsx"
    filepath = os.path.join("uploads", "exports", filename)
    
    # 确保目录存在
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    return filepath


def build_stats_scores_excel(rows: list[dict], output_path: str) -> str:
    """构建统计总览批量导出Excel"""
    wb = _create_workbook()
    ws = wb.active
    ws.title = "统计总览"

    headers = [
        "场次名称",
        "候选人姓名",
        "候选人手机号",
        "评委姓名",
        "总分",
        "评分状态",
        "提交时间",
        "创建时间",
        "更新时间",
    ]
    keys = [
        "session_name",
        "candidate_name",
        "candidate_phone",
        "interviewer_name",
        "total_score",
        "score_status",
        "submitted_at",
        "created_at",
        "updated_at",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    _apply_styles(ws, len(rows) + 1, len(headers))

    directory = os.path.dirname(output_path)
    if directory:
        os.makedirs(directory, exist_ok=True)

    wb.save(output_path)
    return output_path


def build_session_scores_detail_excel(
    rows: list[dict],
    dynamic_columns: list[str],
    output_path: str,
    dimension_columns: list[str] = None,
) -> str:
    """构建统计详情导出Excel（固定列+维度得分动态列+文字反馈动态列）"""
    wb = _create_workbook()
    ws = wb.active
    ws.title = "统计详情"

    base_headers = [
        "场次名称",
        "候选人姓名",
        "候选人手机号",
        "评委姓名",
        "总分",
    ]
    time_headers = [
        "提交时间",
        "创建时间",
        "更新时间",
    ]
    # 列顺序：基础固定列 + 维度得分列 + 时间列 + 文字反馈动态列
    headers = base_headers + (dimension_columns or []) + time_headers + dynamic_columns

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    _apply_styles(ws, len(rows) + 1, len(headers))

    directory = os.path.dirname(output_path)
    if directory:
        os.makedirs(directory, exist_ok=True)

    wb.save(output_path)
    return output_path


def _create_workbook() -> Workbook:
    """创建Excel工作簿"""
    wb = Workbook()
    return wb


def _set_column_width(ws, headers: list, include_feedbacks: bool = False):
    """设置列宽"""
    # 基础列宽
    column_widths = [8, 12, 15, 10]  # 序号、姓名、手机号、平均分
    
    # 评委列宽
    interviewer_count = len(headers) - 4 - (1 if include_feedbacks else 0)
    column_widths.extend([10] * interviewer_count)
    
    # 评语列宽
    if include_feedbacks:
        column_widths.append(40)
    
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_styles(ws, row_count: int, col_count: int):
    """应用样式（表头加粗、边框等）"""
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用边框和对齐
    for row in range(1, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # 数据行居中对齐（除了评语列）
            if row > 1:
                if col < col_count or col_count <= 4:  # 不是最后一列或没有评语列
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # 评语列左对齐
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
